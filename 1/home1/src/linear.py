import numpy as np
import openpyxl
import matplotlib.pyplot as plt

__all__ = ('sheet_to_array', 'ab', 'save_pict', 'kill_extend')

def sheet_to_array(filepath, sheetname=None):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    if sheetname is None:
        sheetname = wb.sheetnames[0]
    sh = wb[sheetname]
    r = sh.max_row - sh.min_row + 1
    c = sh.max_column - sh.min_column + 1
    result = np.zeros((r,c),dtype=np.float64)
    for r,row in enumerate(sh.rows):
        for c, cell in enumerate(row):
            result[r,c] = cell.value
    return result

def ab(D):
    # X = D[:,0]
    # Y = D[:,1]
    N = len(D[:,0])
    Sxy = np.sum(np.prod(D, axis=1))    #сумма произведений x * y
    #сумма всех x
    S = np.sum(D, axis=0)
    Sxx = np.sum(D[:,0]**2)
    a = (N*Sxy - S[0]*S[1]) / (N * Sxx - S[0]**2)
    b = (S[1] - a * S[0]) / N
    return a, b


def save_pict(Dt, figpath, *, a=None, b=None, add_points=None):
    # Dt = sheet_to_array(filepath)
    if a is None or b is None:
        a, b = ab(Dt)
    # x1 = min(Dt[:,0])
    # x2 = max(Dt[:,0])
    # y1 = a*x1 + b
    # y2 = a*x2 + b
    x = np.array([min(Dt[:,0]), max(Dt[:,0])], dtype = Dt.dtype)
    y = a*x + b
    plt.plot(Dt[:,0], Dt[:,1], 'ys')
    if add_points is not None:
        plt.plot(add_points[:,0], add_points[:,1], 'ro')
    plt.plot(x,y,'g--')
    plt.savefig(figpath)
    plt.clf()


#сделать для удаления нескольких точек
def kill_extend(Dt, *, a=None, b=None):
    if a is None or b is None:
        a, b = ab(Dt)

    X = Dt[:,0]
    Y = a*X + b
    deltas = (Dt[:,1] - Y)**2
    k = np.argmax(deltas)
    return (np.delete(Dt, k, axis=0), np.array([Dt[k, :]], dtype=Dt.dtype))

